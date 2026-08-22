import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import unicodedata
import shutil
import os
import hashlib
import hmac
import secrets

ARQUIVO_EXCEL = "BM_MOTOS_Gestao_Estoque_Caixa.xlsx"
TEMPLATE_EXCEL = "BM_MOTOS_Gestao_Estoque_Caixa_Template.xlsx"

# Se o arquivo original não existir (como acontece no GitHub/Streamlit Cloud),
# cria uma cópia a partir do template para o app poder rodar.
if not os.path.exists(ARQUIVO_EXCEL):
    if os.path.exists(TEMPLATE_EXCEL):
        shutil.copy(TEMPLATE_EXCEL, ARQUIVO_EXCEL)
    else:
        st.error("Erro: Nem a planilha original nem o template foram encontrados.")
        st.stop()



# ============================================================
# CONFIGURAÇÕES GERAIS
# ============================================================


try:
    SENHA_ADMIN = str(st.secrets["SENHA_ADMIN"]).strip()
except Exception:
    SENHA_ADMIN = os.getenv("SENHA_ADMIN", "").strip()


st.set_page_config(
    page_title="BM MOTOS",
    page_icon="🏍️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def remover_acentos(texto):
    """Remove acentos, espaços extras e padroniza para busca"""
    if pd.isna(texto) or texto is None:
        return ""
    texto = str(texto).strip().lower()
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join([c for c in nfkd if not unicodedata.combining(c)])

def carregar_dados():
    """Carrega todas as abas com tratamento de tipos"""
    abas = {}
    for aba in ["Produtos", "Vendas", "Itens_Venda", "Movimentacoes_Estoque", "Mecanicos", "Caixa"]:
        df = pd.read_excel(ARQUIVO_EXCEL, sheet_name=aba)
        abas[aba] = df
    return abas

def limpar_cache():
    """Limpa cache de dados após salvar"""
    st.cache_data.clear()

def gerar_id_venda(df_vendas):
    """Gera ID de venda sequencial seguro"""
    if df_vendas.empty or "ID_Venda" not in df_vendas.columns:
        return 1
    return int(df_vendas["ID_Venda"].max()) + 1

def validar_estoque_suficiente(df_produtos, lista_itens):
    """Verifica se tem estoque antes de permitir venda"""
    erros = []
    for item in lista_itens:
        filtro = df_produtos["ID_Produto"] == item["ID_Produto"]
        if not filtro.any():
            erros.append(f"❌ Produto ID {item['ID_Produto']} não encontrado")
            continue
        estoque_atual = int(df_produtos.loc[filtro, "Quantidade"].iloc[0])
        if item["Quantidade"] > estoque_atual:
            erros.append(
                f"❌ {item['Nome_Peca']}: Solicitado {item['Quantidade']}, disponível {estoque_atual}"
            )
    return erros

# ============================================================
# SALVAR VENDA COMPLETA
# ============================================================

def salvar_venda_completa(dados_venda, itens_venda, movimentacoes_estoque, movimentacao_caixa):
    wb = load_workbook(ARQUIVO_EXCEL)

    # Vendas
    ws_vendas = wb["Vendas"]
    ws_vendas.append([
        dados_venda["ID_Venda"],
        dados_venda["Data_Hora"],
        dados_venda["Mecanico"],
        round(dados_venda["Valor_Pecas"], 2),
        round(dados_venda["Valor_Mao_Obra"], 2),
        round(dados_venda["Valor_Total"], 2),
        dados_venda["Observacoes"]
    ])

    # Itens da Venda
    ws_itens = wb["Itens_Venda"]
    for item in itens_venda:
        ws_itens.append([
            item["ID_Venda"],
            item["ID_Produto"],
            item["Nome_Peca"],
            item["Quantidade"],
            round(item["Valor_Unitario"], 2),
            round(item["Subtotal"], 2)
        ])

    # Estoque e Movimentações
    ws_produtos = wb["Produtos"]
    ws_mov_estoque = wb["Movimentacoes_Estoque"]
    ultima_linha_mov = 2 if ws_mov_estoque.max_row < 2 else ws_mov_estoque.max_row + 1

    for mov in movimentacoes_estoque:
        encontrado = False
        for linha in range(2, ws_produtos.max_row + 1):
            id_planilha = ws_produtos.cell(row=linha, column=1).value
            if id_planilha == mov["ID_Produto"]:
                estoque_anterior = int(ws_produtos.cell(row=linha, column=6).value or 0)
                estoque_atual = estoque_anterior - mov["Quantidade"]
                ws_produtos.cell(row=linha, column=6).value = estoque_atual
                ws_mov_estoque.append([
                    ultima_linha_mov - 1,
                    mov["Data_Hora"],
                    "Saída",
                    mov["ID_Produto"],
                    mov["Nome_Peca"],
                    mov["Quantidade"],
                    estoque_anterior,
                    estoque_atual,
                    mov["ID_Venda"]
                ])
                ultima_linha_mov += 1
                encontrado = True
                break
        if not encontrado:
            raise ValueError(f"Produto ID {mov['ID_Produto']} não encontrado na planilha")

    # Caixa
    ws_caixa = wb["Caixa"]
    proximo_id_caixa = 1 if ws_caixa.max_row < 2 else ws_caixa.max_row
    ws_caixa.append([
        proximo_id_caixa,
        movimentacao_caixa["Data_Hora"],
        movimentacao_caixa["Tipo"],
        movimentacao_caixa["Categoria"],
        movimentacao_caixa["Descricao"],
        round(movimentacao_caixa["Valor"], 2),
        movimentacao_caixa["ID_Venda"]
    ])

    wb.save(ARQUIVO_EXCEL)
    limpar_cache()

# ============================================================
# ENTRADA DE ESTOQUE
# ============================================================

def registrar_entrada_estoque(id_produto, nome_peca, quantidade):
    if quantidade <= 0:
        raise ValueError("Quantidade deve ser maior que zero")

    wb = load_workbook(ARQUIVO_EXCEL)
    ws_produtos = wb["Produtos"]
    ws_mov_estoque = wb["Movimentacoes_Estoque"]
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    proximo_id_mov = 1 if ws_mov_estoque.max_row < 2 else ws_mov_estoque.max_row

    encontrado = False
    for linha in range(2, ws_produtos.max_row + 1):
        if ws_produtos.cell(row=linha, column=1).value == id_produto:
            estoque_anterior = int(ws_produtos.cell(row=linha, column=6).value or 0)
            estoque_atual = estoque_anterior + quantidade
            ws_produtos.cell(row=linha, column=6).value = estoque_atual
            ws_mov_estoque.append([
                proximo_id_mov,
                data_hora,
                "Entrada",
                id_produto,
                nome_peca,
                quantidade,
                estoque_anterior,
                estoque_atual,
                None
            ])
            encontrado = True
            novo_valor = estoque_atual
            break

    if not encontrado:
        wb.close()
        raise ValueError(f"Produto '{nome_peca}' não encontrado")

    wb.save(ARQUIVO_EXCEL)
    limpar_cache()
    return novo_valor

# ============================================================
# CADASTRAR PRODUTO NOVO
# ============================================================

def cadastrar_novo_produto(nome, categoria, preco_custo, preco_venda, quantidade, unidade):
    if not nome.strip():
        raise ValueError("Nome do produto é obrigatório")
    if preco_venda <= 0:
        raise ValueError("Preço de venda deve ser maior que zero")
    if quantidade <= 0:
        raise ValueError("Quantidade inicial deve ser maior que zero")

    wb = load_workbook(ARQUIVO_EXCEL)
    ws_produtos = wb["Produtos"]
    ws_mov_estoque = wb["Movimentacoes_Estoque"]

    novo_id = 1 if ws_produtos.max_row < 2 else ws_produtos.max_row
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    proximo_id_mov = 1 if ws_mov_estoque.max_row < 2 else ws_mov_estoque.max_row

    ws_produtos.append([novo_id, nome.strip(), categoria, preco_custo, preco_venda, quantidade, unidade])
    ws_mov_estoque.append([
        proximo_id_mov, data_hora, "Entrada", novo_id, nome.strip(), quantidade, 0, quantidade, None
    ])

    wb.save(ARQUIVO_EXCEL)
    limpar_cache()
    return novo_id

# ============================================================
# LOGIN
# ============================================================

def gerar_hash_senha(senha, salt=None, iteracoes=200_000):
    """Gera um hash PBKDF2 com salt para armazenar senhas sem texto aberto."""
    senha_bytes = str(senha).encode("utf-8")
    salt = salt or secrets.token_hex(16)
    hash_bytes = hashlib.pbkdf2_hmac(
        "sha256", senha_bytes, salt.encode("utf-8"), iteracoes
    )
    return f"pbkdf2_sha256${iteracoes}${salt}${hash_bytes.hex()}"


def verificar_hash_senha(senha_digitada, senha_armazenada):
    """Compara uma senha digitada com o hash PBKDF2 armazenado na planilha."""
    try:
        algoritmo, iteracoes_texto, salt, hash_esperado = str(senha_armazenada).strip().split("$", 3)
        if algoritmo != "pbkdf2_sha256":
            return False
        iteracoes = int(iteracoes_texto)
        hash_calculado = hashlib.pbkdf2_hmac(
            "sha256",
            str(senha_digitada).encode("utf-8"),
            salt.encode("utf-8"),
            iteracoes,
        ).hex()
        return hmac.compare_digest(hash_calculado, hash_esperado)
    except (TypeError, ValueError):
        return False


def fazer_login_mecanico(nome_selecionado, senha_digitada, df_mecanicos):
    linha = df_mecanicos[df_mecanicos["Nome"] == nome_selecionado]
    if linha.empty:
        return False
    senha_armazenada = linha.iloc[0]["Senha"]
    return verificar_hash_senha(senha_digitada, senha_armazenada)

def fazer_login_admin(senha):
    return str(senha).strip() == SENHA_ADMIN

def fazer_logout():
    for chave in list(st.session_state.keys()):
        del st.session_state[chave]
    st.rerun()

# Inicializar sessão
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
    st.session_state.nivel_acesso = None
    st.session_state.usuario = None
if "carrinho" not in st.session_state:
    st.session_state.carrinho = []

# ============================================================
# TELA DE LOGIN
# ============================================================

if not st.session_state.autenticado:
    st.title("🏍️ BM MOTOS")
    st.markdown("---")

    _, centro, _ = st.columns([1, 1, 1])
    with centro:
        st.subheader("🔐 Acesso ao Sistema")

        try:
            dados_login = carregar_dados()
            df_mec = dados_login["Mecanicos"]
            lista_nomes = sorted(df_mec["Nome"].dropna().unique().tolist())
        except Exception as e:
            st.error(f"❌ Erro ao carregar dados: {e}")
            st.stop()

        tipo = st.radio("Quem é você?", ["👨‍🔧 Mecânico", "👨‍💼 Administrador"], horizontal=True)

        if tipo == "👨‍🔧 Mecânico":
            nome = st.selectbox("Selecione seu nome", [""] + lista_nomes, placeholder="Escolha seu nome...")
            senha = st.text_input("Digite sua senha", type="password")
            if st.button("Entrar", type="primary", use_container_width=True):
                if not nome:
                    st.warning("⚠️ Selecione seu nome")
                elif fazer_login_mecanico(nome, senha, df_mec):
                    st.session_state.autenticado = True
                    st.session_state.nivel_acesso = "mecanico"
                    st.session_state.usuario = nome
                    st.success(f"✅ Bem-vindo, {nome}!")
                    st.rerun()
                else:
                    st.error("❌ Senha incorreta!")
        else:
            senha = st.text_input("Senha do Administrador", type="password")
            if st.button("Entrar como Admin", type="primary", use_container_width=True):
                if fazer_login_admin(senha):
                    st.session_state.autenticado = True
                    st.session_state.nivel_acesso = "admin"
                    st.session_state.usuario = "Administrador"
                    st.success("✅ Acesso liberado!")
                    st.rerun()
                else:
                    st.error("❌ Senha incorreta!")

        st.markdown("---")
        st.info("💡 Selecione seu nome e digite sua senha. Mantenha sua senha em sigilo.")
    st.stop()

# ============================================================
# ÁREA PRINCIPAL
# ============================================================

st.title("🏍️ BM MOTOS")

cab1, cab2, cab3 = st.columns([4, 2, 1])
with cab2:
    if st.session_state.nivel_acesso == "admin":
        st.success("👨‍💼 Administrador")
    else:
        st.info(f"👨‍🔧 {st.session_state.usuario}")
with cab3:
    if st.button("🚪 Sair", use_container_width=True):
        fazer_logout()

st.markdown("---")

# Carregar dados
dados = carregar_dados()
df_produtos = dados["Produtos"]
df_mecanicos = dados["Mecanicos"]
df_vendas = dados["Vendas"]

# Definir abas
if st.session_state.nivel_acesso == "admin":
    aba_venda, aba_estoque_entrada, aba_consulta, aba_historico = st.tabs([
        "🛒 Nova Venda", "📥 Entrada de Estoque", "📦 Consultar Estoque", "📋 Vendas Realizadas"
    ])
else:
    aba_venda = st.container()  # Apenas venda para mecânico

# ============================================================
# ABA NOVA VENDA
# ============================================================

with aba_venda:
    st.subheader("🛒 Nova Venda")

    # Dados da venda
    st.markdown("#### 👤 Dados da Venda")
    col1, col2 = st.columns(2)
    with col1:
        if st.session_state.nivel_acesso == "admin":
            nome_mecanico_venda = st.selectbox(
                "Mecânico responsável",
                ["BM MOTOS"] + sorted(df_mecanicos["Nome"].dropna().tolist())
            )
        else:
            nome_mecanico_venda = st.session_state.usuario
            st.info(f"👨‍🔧 Responsável: **{nome_mecanico_venda}**")
    with col2:
        valor_mao_obra = st.number_input(
            "🔧 Valor da Mão de Obra (R$)",
            min_value=0.0, step=5.0, format="%.2f"
        )
    obs = st.text_input("📝 Observações (opcional)", placeholder="Ex: Troca de óleo, revisão de freios...")
    st.markdown("---")

    # Busca de peças - RESULTADOS APARECEM CONFORME DIGITA
    st.markdown("#### 🔍 Pesquisar Peças")
    termo_busca = st.text_input(
        "Digite o nome da peça",
        placeholder="Ex.: óleo, filtro, vela, freio...",
        help="Não precisa usar acentos! Resultados aparecem automaticamente",
        key="busca_venda",
        on_change=None
    )

    # Filtrar com busca inteligente - SEM PRECISAR APERTAR ENTER
    if termo_busca:
        termo = remover_acentos(termo_busca)
        filtro = df_produtos["Nome"].apply(lambda x: termo in remover_acentos(x))
        produtos_filtrados = df_produtos[filtro & (df_produtos["Quantidade"] > 0)]
        if len(produtos_filtrados):
            st.success(f"✅ {len(produtos_filtrados)} peça(s) encontrada(s)")
        else:
            st.warning("⚠️ Nenhuma peça disponível com esse nome")
    else:
        produtos_filtrados = df_produtos[df_produtos["Quantidade"] > 0]

    # Selecionar peça
    if not produtos_filtrados.empty:
        col_p, col_q, col_b = st.columns([3, 1, 1])
        with col_p:
            peca_nome = st.selectbox("Peça", produtos_filtrados["Nome"].tolist())
        dados_peca = produtos_filtrados[produtos_filtrados["Nome"] == peca_nome].iloc[0]
        estoque_disp = int(dados_peca["Quantidade"])
        preco_unit = round(float(dados_peca["Preco_Venda"]), 2)
        with col_q:
            qtd_escolhida = st.number_input("Quantidade", min_value=1, max_value=estoque_disp, step=1)
        with col_b:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("➕ Adicionar", use_container_width=True):
                st.session_state.carrinho.append({
                    "ID_Produto": int(dados_peca["ID_Produto"]),
                    "Nome_Peca": peca_nome,
                    "Quantidade": qtd_escolhida,
                    "Valor_Unitario": preco_unit,
                    "Subtotal": round(qtd_escolhida * preco_unit, 2)
                })
                st.toast(f"✅ {qtd_escolhida}x {peca_nome} adicionado ao carrinho!", icon="✅")
                st.rerun()

        st.info(f"📦 Estoque disponível: {estoque_disp} | 💰 R$ {preco_unit:.2f}")

    st.markdown("---")

    # Carrinho
    st.markdown("#### 🛒 Carrinho")
    if st.session_state.carrinho:
        df_carr = pd.DataFrame(st.session_state.carrinho)
        st.dataframe(df_carr, use_container_width=True, hide_index=True)

        total_pecas = round(df_carr["Subtotal"].sum(), 2)
        total_geral = round(total_pecas + valor_mao_obra, 2)

        r1, r2, r3 = st.columns(3)
        r1.metric("Peças", f"R$ {total_pecas:.2f}")
        r2.metric("Mão de Obra", f"R$ {valor_mao_obra:.2f}")
        r3.metric("TOTAL", f"R$ {total_geral:.2f}")

        cl, cf = st.columns([1, 2])
        with cl:
            if st.button("🗑️ Limpar", use_container_width=True):
                st.session_state.carrinho = []
                st.toast("🗑️ Carrinho limpo!", icon="🗑️")
                st.rerun()
        with cf:
            if st.button("✅ FINALIZAR VENDA", type="primary", use_container_width=True):
                erros = validar_estoque_suficiente(df_produtos, st.session_state.carrinho)
                if erros:
                    for e in erros:
                        st.error(e)
                    st.stop()

                id_venda = gerar_id_venda(df_vendas)
                data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                dados_venda = {
                    "ID_Venda": id_venda,
                    "Data_Hora": data_hora,
                    "Mecanico": nome_mecanico_venda,
                    "Valor_Pecas": total_pecas,
                    "Valor_Mao_Obra": valor_mao_obra,
                    "Valor_Total": total_geral,
                    "Observacoes": obs
                }

                itens_venda = []
                mov_estoque = []
                for item in st.session_state.carrinho:
                    item["ID_Venda"] = id_venda
                    itens_venda.append(item)
                    mov_estoque.append({
                        "Data_Hora": data_hora,
                        "ID_Produto": item["ID_Produto"],
                        "Nome_Peca": item["Nome_Peca"],
                        "Quantidade": item["Quantidade"],
                        "ID_Venda": id_venda
                    })

                mov_caixa = {
                    "Data_Hora": data_hora,
                    "Tipo": "Entrada",
                    "Categoria": "Venda de Peças e Serviços",
                    "Descricao": f"Venda #{id_venda} - {nome_mecanico_venda}",
                    "Valor": total_geral,
                    "ID_Venda": id_venda
                }

                try:
                    salvar_venda_completa(dados_venda, itens_venda, mov_estoque, mov_caixa)
                    st.session_state.carrinho = []
                    st.toast(f"🎉 Venda #{id_venda} registrada com SUCESSO!", icon="✅")
                    st.success(f"""
                    ✅ Venda #{id_venda} finalizada!
                    📅 {data_hora}
                    👤 Responsável: {nome_mecanico_venda}
                    💰 Total: R$ {total_geral:.2f}
                    """)
                    st.balloons()
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Erro ao salvar: {e}")
    else:
        st.info("ℹ️ Adicione peças acima para montar a venda")

# ============================================================
# ABA ENTRADA DE ESTOQUE (ADMIN)
# ============================================================

if st.session_state.nivel_acesso == "admin":
    with aba_estoque_entrada:
        st.subheader("📥 Entrada de Estoque")
        tab_existente, tab_novo = st.tabs(["➕ Produto Existente", "🆕 Produto Novo"])

        with tab_existente:
            b_usuario = st.text_input("Pesquisar produto", placeholder="Ex.: óleo, filtro...", key="busca_estoque_entrada")
            if b_usuario:
                t = remover_acentos(b_usuario)
                filtro = df_produtos["Nome"].apply(lambda x: t in remover_acentos(x))
                lista_prod = df_produtos[filtro]
            else:
                lista_prod = df_produtos.copy()

            if lista_prod.empty:
                st.warning("Nenhum produto encontrado")
            else:
                sel_nome = st.selectbox("Selecione o produto", lista_prod["Nome"].tolist(), key="sel_prod_entrada")
                sel_dados = lista_prod[lista_prod["Nome"] == sel_nome].iloc[0]
                est_atual = int(sel_dados["Quantidade"])
                st.info(f"📦 Estoque atual: {est_atual} {sel_dados['Unidade']}")
                qtd_add = st.number_input("Quantidade a adicionar", min_value=1, step=1, key="qtd_add_entrada")
                if st.button("✅ Confirmar Entrada", type="primary", use_container_width=True):
                    try:
                        novo_est = registrar_entrada_estoque(int(sel_dados["ID_Produto"]), sel_nome, qtd_add)
                        st.toast(f"✅ Estoque atualizado! {est_atual} → {novo_est}", icon="📦")
                        st.success(f"""
                        ✅ Entrada registrada com sucesso!
                        📦 Produto: {sel_nome}
                        ➕ Adicionado: {qtd_add} unidades
                        📊 Estoque: {est_atual} → {novo_est}
                        """)
                        st.balloons()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ {e}")

        with tab_novo:
            st.markdown("#### Cadastrar Produto Novo")
            c1, c2 = st.columns(2)
            with c1:
                nv_nome = st.text_input("Nome do Produto *", key="nv_nome")
                nv_cat = st.selectbox("Categoria", [
                    "Óleos e Lubrificantes", "Filtros", "Freios", "Elétrica",
                    "Transmissão", "Cabos", "Ignição", "Pneus", "Suspensão", "Outros"
                ], key="nv_cat")
                nv_und = st.selectbox("Unidade", ["Un", "Jogo", "Par", "Litro", "Kg", "Metro"], key="nv_und")
            with c2:
                nv_custo = st.number_input("Preço de Custo (R$)", min_value=0.0, step=0.50, format="%.2f", key="nv_custo")
                nv_venda = st.number_input("Preço de Venda (R$) *", min_value=0.01, step=0.50, format="%.2f", key="nv_venda")
                nv_qtd = st.number_input("Quantidade Inicial *", min_value=1, step=1, key="nv_qtd")

            if st.button("🆕 Cadastrar", type="primary", use_container_width=True):
                try:
                    novo_id = cadastrar_novo_produto(nv_nome, nv_cat, nv_custo, nv_venda, nv_qtd, nv_und)
                    st.toast(f"✅ Produto cadastrado! ID: {novo_id}", icon="🆕")
                    st.success(f"""
                    ✅ Produto cadastrado com sucesso!
                    🆔 ID: {novo_id}
                    📌 Nome: {nv_nome}
                    📦 Quantidade: {nv_qtd} {nv_und}
                    """)
                    st.balloons()
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ {e}")

# ============================================================
# ABA CONSULTA DE ESTOQUE (ADMIN)
# ============================================================

if st.session_state.nivel_acesso == "admin":
    with aba_consulta:
        st.subheader("📦 Consulta de Estoque")
        b_estoque = st.text_input("Pesquisar", placeholder="Nome do produto...", key="busca_consulta_estoque")
        if b_estoque:
            t = remover_acentos(b_estoque)
            filtro = df_produtos["Nome"].apply(lambda x: t in remover_acentos(x))
            df_consulta = df_produtos[filtro].copy()
        else:
            df_consulta = df_produtos.copy()

        df_consulta["Status"] = df_consulta["Quantidade"].apply(
            lambda x: "🟢 Normal" if x > 10 else ("🟡 Baixo" if x > 0 else "🔴 Esgotado")
        )
        st.dataframe(
            df_consulta[["ID_Produto", "Nome", "Categoria", "Preco_Custo", "Preco_Venda", "Quantidade", "Status"]],
            use_container_width=True, hide_index=True
        )

        t1, t2, t3 = st.columns(3)
        t1.metric("Total Produtos", len(df_produtos))
        t2.metric("⚠️ Estoque Baixo", len(df_produtos[(df_produtos["Quantidade"] > 0) & (df_produtos["Quantidade"] <= 10)]))
        t3.metric("❌ Esgotados", len(df_produtos[df_produtos["Quantidade"] == 0]))

# ============================================================
# ABA HISTÓRICO DE VENDAS (ADMIN)
# ============================================================

if st.session_state.nivel_acesso == "admin":
    with aba_historico:
        st.subheader("📋 Histórico de Vendas")
        if df_vendas.empty:
            st.info("Nenhuma venda registrada ainda")
        else:
            df_v = df_vendas.sort_values("ID_Venda", ascending=False)
            st.dataframe(df_v, use_container_width=True, hide_index=True)
            t_vendas = len(df_vendas)
            faturamento = round(df_vendas["Valor_Total"].sum(), 2)
            ticket = round(faturamento / t_vendas, 2) if t_vendas else 0
            v1, v2, v3 = st.columns(3)
            v1.metric("Total Vendas", t_vendas)
            v2.metric("Faturamento", f"R$ {faturamento:.2f}")
            v3.metric("Ticket Médio", f"R$ {ticket:.2f}")

# ============================================================
# RODAPÉ
# ============================================================
st.markdown("---")
st.caption("🏍️ BM MOTOS - Sistema de Gestão de Oficina | v2.1 Atualizado")